#!/usr/bin/env python3
"""
Zabbix SLA Report Generator

Generates Excel reports with availability data for specified host groups.
Supports custom date ranges, previous week, and previous month reporting.
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ZabbixAPI:
    """Zabbix API wrapper class."""

    def __init__(self, url: str, token: str):
        self.url = url.rstrip("/") + "/api_jsonrpc.php"
        self.token = token
        self.request_id = 0

    def _call(self, method: str, params: dict = None, use_auth: bool = True) -> Any:
        """Make an API call to Zabbix."""
        self.request_id += 1
        payload = {
            "jsonrpc": "2.0",
            "method": method,
            "params": params or {},
            "id": self.request_id,
        }

        headers = {"Content-Type": "application/json"}
        if use_auth:
            headers["Authorization"] = f"Bearer {self.token}"

        try:
            response = requests.post(self.url, json=payload, headers=headers, timeout=60)
            response.raise_for_status()
            result = response.json()

            if "error" in result:
                raise Exception(f"Zabbix API error: {result['error']}")

            return result.get("result")
        except requests.exceptions.RequestException as e:
            raise Exception(f"Request failed: {e}")

    def get_host_groups(self, names: list[str] = None) -> list[dict]:
        """Get host groups, optionally filtered by name."""
        params = {"output": ["groupid", "name"]}
        if names:
            params["filter"] = {"name": names}
        return self._call("hostgroup.get", params)

    def get_hosts_in_group(self, group_id: str) -> list[dict]:
        """Get all hosts in a specific host group."""
        params = {
            "output": ["hostid", "host", "name"],
            "groupids": group_id,
            "filter": {"status": 0},  # Only enabled hosts
        }
        return self._call("host.get", params)

    def get_problems(
        self, host_ids: list[str], time_from: int, time_till: int
    ) -> list[dict]:
        """Get problems/events for hosts in a time range."""
        params = {
            "output": ["eventid", "objectid", "clock", "r_clock", "severity"],
            "hostids": host_ids,
            "time_from": time_from,
            "time_till": time_till,
            "source": 0,  # Triggers
            "object": 0,  # Trigger events
            "severities": [4],  # High severity only (ICMP unavailable)
            "selectHosts": ["hostid", "host", "name"],
            "sortfield": ["clock"],
            "sortorder": "ASC",
        }
        return self._call("event.get", params)

    def get_triggers(self, host_ids: list[str]) -> list[dict]:
        """Get triggers for hosts (for availability calculation)."""
        params = {
            "output": ["triggerid", "description", "priority"],
            "hostids": host_ids,
            "filter": {"status": 0},  # Enabled triggers only
            "min_severity": 3,  # Average and above
            "selectHosts": ["hostid", "host", "name"],
        }
        return self._call("trigger.get", params)

    def get_sla(self, sla_ids: list[str] = None) -> list[dict]:
        """Get SLA definitions."""
        params = {"output": "extend", "selectServiceTags": "extend"}
        if sla_ids:
            params["slaids"] = sla_ids
        return self._call("sla.get", params)

    def get_host_availability(
        self, host_id: str, time_from: int, time_till: int
    ) -> dict:
        """
        Calculate host availability based on event records.
        Returns availability percentage.
        Only counts 'Unavailable by ICMP ping' problems.

        Uses event.get (not problem.get) because the problem table may not
        retain resolved problems depending on housekeeper/recent settings.
        Two-step: fetch PROBLEM events, then batch-fetch recovery events.
        """
        # Step 1: Get PROBLEM events (value=1) within the time window
        params = {
            "output": ["eventid", "clock", "r_eventid", "name"],
            "hostids": [host_id],
            "time_from": time_from,
            "time_till": time_till,
            "source": 0,   # Triggers
            "object": 0,   # Trigger events
            "value": "1",  # PROBLEM events only
            "sortfield": ["clock"],
            "sortorder": "ASC",
        }
        events_in_window = self._call("event.get", params)

        # Step 2: Get PROBLEM events that started BEFORE the window
        # (they may still have been active during the window)
        params_before = {
            "output": ["eventid", "clock", "r_eventid", "name"],
            "hostids": [host_id],
            "time_till": time_from - 1,
            "source": 0,
            "object": 0,
            "value": "1",
            "sortfield": ["clock"],
            "sortorder": "DESC",
            "limit": 50,
        }
        events_before = self._call("event.get", params_before)

        # Combine all candidate events
        all_events = list(events_in_window) + list(events_before)

        # Step 3: Collect all recovery event IDs so we can batch-fetch their timestamps
        recovery_ids = []
        for evt in all_events:
            r_id = evt.get("r_eventid", "0")
            if r_id and r_id != "0":
                recovery_ids.append(r_id)

        # Step 4: Batch-fetch recovery events to get their clock (= recovery time)
        recovery_map = {}  # r_eventid -> clock
        if recovery_ids:
            recovery_params = {
                "output": ["eventid", "clock"],
                "eventids": recovery_ids,
            }
            recovery_events = self._call("event.get", recovery_params)
            for rev in recovery_events:
                recovery_map[rev["eventid"]] = int(rev["clock"])

        # Step 5: Calculate downtime
        total_seconds = time_till - time_from
        downtime_seconds = 0

        for event in all_events:
            # Only count "Unavailable by ICMP ping" problems
            event_name = event.get("name", "").lower()
            if "unavailable by icmp" not in event_name:
                continue

            event_start = int(event["clock"])
            r_eventid = event.get("r_eventid", "0")

            if r_eventid and r_eventid != "0":
                # Resolved: look up recovery time
                event_end = recovery_map.get(r_eventid, time_till)
            else:
                # Still active / unresolved
                event_end = time_till

            # Clamp to the time range
            actual_start = max(event_start, time_from)
            actual_end = min(event_end, time_till)

            if actual_end > actual_start:
                downtime_seconds += actual_end - actual_start

        if total_seconds > 0:
            availability = ((total_seconds - downtime_seconds) / total_seconds) * 100
        else:
            availability = 100.0

        return {
            "availability": round(availability, 2),
            "downtime_seconds": downtime_seconds,
            "total_seconds": total_seconds,
        }


class DateRangeCalculator:
    """Calculate date ranges for different report types."""

    @staticmethod
    def get_custom_range(start_date: str, end_date: str) -> tuple[datetime, datetime]:
        """Parse custom date range."""
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )
        return start, end

    @staticmethod
    def get_previous_week() -> tuple[datetime, datetime]:
        """Get previous week (Monday to Sunday)."""
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        last_sunday = today - timedelta(days=today.weekday() + 1)
        last_monday = last_sunday - timedelta(days=6)
        return last_monday, last_sunday.replace(hour=23, minute=59, second=59)

    @staticmethod
    def get_previous_month() -> tuple[datetime, datetime]:
        """Get previous month (1st to last day)."""
        today = datetime.now()
        first_of_this_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_of_prev_month = first_of_this_month - timedelta(days=1)
        first_of_prev_month = last_of_prev_month.replace(day=1)
        return first_of_prev_month, last_of_prev_month.replace(
            hour=23, minute=59, second=59
        )

    @staticmethod
    def get_availability_periods(reference_date: datetime = None) -> dict:
        """
        Get the three availability periods:
        - 1 Day: yesterday (last 24 hours)
        - 7 Days: last 7 days
        - Previous Month: actual previous calendar month
        """
        if reference_date is None:
            reference_date = datetime.now()

        today_start = reference_date.replace(hour=0, minute=0, second=0, microsecond=0)

        # 1 Day - Yesterday (last 24 hours)
        day_1_end = today_start - timedelta(seconds=1)
        day_1_start = today_start - timedelta(days=1)

        # 7 Days - Last 7 days (including yesterday)
        day_7_end = today_start - timedelta(seconds=1)
        day_7_start = today_start - timedelta(days=7)

        # Previous Month - Actual calendar month
        first_of_this_month = today_start.replace(day=1)
        last_of_prev_month = first_of_this_month - timedelta(days=1)
        first_of_prev_month = last_of_prev_month.replace(day=1)
        prev_month_start = first_of_prev_month
        prev_month_end = last_of_prev_month.replace(hour=23, minute=59, second=59)

        return {
            "1_day": (day_1_start, day_1_end),
            "7_days": (day_7_start, day_7_end),
            "prev_month": (prev_month_start, prev_month_end),
        }


class ExcelReportGenerator:
    """Generate Excel reports with conditional formatting."""

    def __init__(self, sla_threshold: float, orange_threshold: float):
        self.sla_threshold = sla_threshold
        self.orange_threshold = orange_threshold
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)

        # Define styles
        self.header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        self.orange_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        self.red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        self.green_font = Font(color="006100")
        self.orange_font = Font(color="9C5700")
        self.red_font = Font(color="9C0006")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_align = Alignment(horizontal="center", vertical="center")

    def get_cell_style(self, value: float) -> tuple:
        """Determine cell fill and font based on value and thresholds."""
        if value >= self.sla_threshold:
            return self.green_fill, self.green_font
        elif value >= self.sla_threshold - self.orange_threshold:
            return self.orange_fill, self.orange_font
        else:
            return self.red_fill, self.red_font

    def create_sheet(self, sheet_name: str, data: list[dict], sla_target: float):
        """Create a worksheet with host availability data."""
        # Truncate sheet name to Excel's 31 character limit
        safe_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
        ws = self.workbook.create_sheet(title=safe_name)

        # Headers
        headers = [
            "Host Name",
            "Host",
            "Availability\n1 Day (%)",
            "Availability\n7 Days (%)",
            "Availability\nPrev Month (%)",
            "Device SLA\n(%)",
            f"SLA Target\n({sla_target}%)",
            "SLA Status",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Write data
        for row_idx, host_data in enumerate(data, 2):
            # Host Name
            cell = ws.cell(row=row_idx, column=1, value=host_data.get("name", ""))
            cell.border = self.border

            # Host (technical name)
            cell = ws.cell(row=row_idx, column=2, value=host_data.get("host", ""))
            cell.border = self.border

            # Availability columns
            availability_cols = [
                ("avail_1_day", 3),
                ("avail_7_days", 4),
                ("avail_prev_month", 5),
                ("device_sla", 6),
            ]

            for key, col in availability_cols:
                value = host_data.get(key, 100.0)
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.number_format = "0.00"
                fill, font = self.get_cell_style(value)
                cell.fill = fill
                cell.font = font
                cell.border = self.border
                cell.alignment = self.center_align

            # SLA Target column (shows the target value)
            cell = ws.cell(row=row_idx, column=7, value=sla_target)
            cell.number_format = "0.00"
            cell.border = self.border
            cell.alignment = self.center_align

            # SLA Status
            device_sla = host_data.get("device_sla", 100.0)
            if device_sla >= self.sla_threshold:
                status = "COMPLIANT"
                fill, font = self.green_fill, self.green_font
            elif device_sla >= self.sla_threshold - self.orange_threshold:
                status = "WARNING"
                fill, font = self.orange_fill, self.orange_font
            else:
                status = "BREACH"
                fill, font = self.red_fill, self.red_font

            cell = ws.cell(row=row_idx, column=8, value=status)
            cell.fill = fill
            cell.font = font
            cell.border = self.border
            cell.alignment = self.center_align

        # Add Overall SLA row at the bottom
        if data:
            overall_row = len(data) + 3  # Skip one row for spacing

            # Calculate overall SLA based on TOTAL TIME (not average)
            total_downtime_1_day = sum(h.get("downtime_1_day", 0) for h in data)
            total_downtime_7_days = sum(h.get("downtime_7_days", 0) for h in data)
            total_downtime_prev_month = sum(h.get("downtime_prev_month", 0) for h in data)

            total_possible_1_day = sum(h.get("total_1_day", 86400) for h in data)
            total_possible_7_days = sum(h.get("total_7_days", 604800) for h in data)
            total_possible_prev_month = sum(h.get("total_prev_month", 2592000) for h in data)

            overall_1_day = ((total_possible_1_day - total_downtime_1_day) / total_possible_1_day * 100) if total_possible_1_day > 0 else 100.0
            overall_7_days = ((total_possible_7_days - total_downtime_7_days) / total_possible_7_days * 100) if total_possible_7_days > 0 else 100.0
            overall_prev_month = ((total_possible_prev_month - total_downtime_prev_month) / total_possible_prev_month * 100) if total_possible_prev_month > 0 else 100.0
            overall_sla = overall_prev_month  # Use prev month for device sheet overall

            # Overall label
            cell = ws.cell(row=overall_row, column=1, value="OVERALL GROUP SLA")
            cell.font = Font(bold=True, size=11)
            cell.border = self.border
            ws.merge_cells(start_row=overall_row, start_column=1, end_row=overall_row, end_column=2)

            # Overall availability values
            overall_cols = [
                (overall_1_day, 3),
                (overall_7_days, 4),
                (overall_prev_month, 5),
                (overall_sla, 6),
            ]

            for value, col in overall_cols:
                cell = ws.cell(row=overall_row, column=col, value=round(value, 2))
                cell.number_format = "0.00"
                fill, font = self.get_cell_style(value)
                cell.fill = fill
                cell.font = Font(bold=True, color=font.color)
                cell.border = self.border
                cell.alignment = self.center_align

            # SLA Target
            cell = ws.cell(row=overall_row, column=7, value=sla_target)
            cell.number_format = "0.00"
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = self.center_align

            # Overall status
            if overall_sla >= self.sla_threshold:
                status = "COMPLIANT"
                fill, font_style = self.green_fill, self.green_font
            elif overall_sla >= self.sla_threshold - self.orange_threshold:
                status = "WARNING"
                fill, font_style = self.orange_fill, self.orange_font
            else:
                status = "BREACH"
                fill, font_style = self.red_fill, self.red_font

            cell = ws.cell(row=overall_row, column=8, value=status)
            cell.fill = fill
            cell.font = Font(bold=True, color=font_style.color)
            cell.border = self.border
            cell.alignment = self.center_align

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(data) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # Handle multi-line headers
                    lines = str(cell_value).split("\n")
                    max_line_length = max(len(line) for line in lines)
                    if max_line_length > max_length:
                        max_length = max_line_length
            ws.column_dimensions[column_letter].width = max_length + 2

        # Set row height for header
        ws.row_dimensions[1].height = 40

    def add_summary_sheet(self, group_summaries: list[dict]):
        """Add a summary sheet with overall SLA compliance."""
        ws = self.workbook.create_sheet(title="Summary", index=0)

        headers = [
            "Host Group",
            "SLA Target\n(%)",
            "Total Hosts",
            "Compliant",
            "Warning",
            "Breach",
            "Overall SLA\n1 Day (%)",
            "Overall SLA\n7 Days (%)",
            "Overall SLA\nPrev Month (%)",
            "Overall\nGroup SLA (%)",
            "SLA Status",
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set header row height
        ws.row_dimensions[1].height = 40

        # Write summary data
        for row_idx, summary in enumerate(group_summaries, 2):
            ws.cell(row=row_idx, column=1, value=summary["group_name"]).border = self.border

            # SLA Target for this group
            group_sla_target = summary.get("sla_threshold", self.sla_threshold)
            cell = ws.cell(row=row_idx, column=2, value=group_sla_target)
            cell.number_format = "0.00"
            cell.border = self.border
            cell.alignment = self.center_align

            cell = ws.cell(row=row_idx, column=3, value=summary["total"])
            cell.border = self.border
            cell.alignment = self.center_align

            cell = ws.cell(row=row_idx, column=4, value=summary["compliant"])
            cell.border = self.border
            cell.alignment = self.center_align

            cell = ws.cell(row=row_idx, column=5, value=summary["warning"])
            cell.border = self.border
            cell.alignment = self.center_align

            cell = ws.cell(row=row_idx, column=6, value=summary["breach"])
            cell.border = self.border
            cell.alignment = self.center_align

            # Overall SLA columns with color coding (use group-specific threshold)
            sla_cols = [
                ("overall_1_day", 7),
                ("overall_7_days", 8),
                ("overall_prev_month", 9),
                ("overall_sla", 10),
            ]

            for key, col in sla_cols:
                value = summary.get(key, 100.0)
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.number_format = "0.00"
                # Use group-specific SLA threshold for color coding
                if value >= group_sla_target:
                    fill, font = self.green_fill, self.green_font
                elif value >= group_sla_target - self.orange_threshold:
                    fill, font = self.orange_fill, self.orange_font
                else:
                    fill, font = self.red_fill, self.red_font
                cell.fill = fill
                cell.font = font
                cell.border = self.border
                cell.alignment = self.center_align

            # SLA Status (use group-specific threshold)
            overall_sla = summary.get("overall_sla", 100.0)
            if overall_sla >= group_sla_target:
                status = "COMPLIANT"
                fill, font = self.green_fill, self.green_font
            elif overall_sla >= group_sla_target - self.orange_threshold:
                status = "WARNING"
                fill, font = self.orange_fill, self.orange_font
            else:
                status = "BREACH"
                fill, font = self.red_fill, self.red_font

            cell = ws.cell(row=row_idx, column=11, value=status)
            cell.fill = fill
            cell.font = font
            cell.border = self.border
            cell.alignment = self.center_align

        # Auto-adjust column widths
        col_widths = [20, 12, 12, 12, 12, 12, 15, 18, 18, 15, 12]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def save(self, filename: str):
        """Save the workbook to a file."""
        self.workbook.save(filename)
        print(f"Report saved to: {filename}")


def load_config(config_path: str) -> dict:
    """Load configuration from YAML file."""
    with open(config_path, "r") as f:
        return yaml.safe_load(f)


def main():
    parser = argparse.ArgumentParser(
        description="Generate Zabbix SLA Reports",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate SLA report based on last 1 day
  python zabbix_sla_report.py --period day

  # Generate SLA report based on last 7 days
  python zabbix_sla_report.py --period week

  # Generate SLA report based on previous month (default)
  python zabbix_sla_report.py --period month

  # Use specific config file
  python zabbix_sla_report.py --config /path/to/config.yaml --period month
        """,
    )

    parser.add_argument(
        "--config",
        "-c",
        default="config.yaml",
        help="Path to configuration file (default: config.yaml)",
    )
    parser.add_argument(
        "--period",
        "-p",
        choices=["day", "week", "month"],
        default="month",
        help="SLA period: day (1 day), week (7 days), or month (previous month)",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Output file path (optional, auto-generated if not specified)",
    )
    parser.add_argument(
        "--groups",
        "-g",
        nargs="+",
        help="Override host groups from config (space-separated)",
    )

    args = parser.parse_args()

    # Load configuration
    config_path = Path(args.config)
    if not config_path.exists():
        print(f"Error: Configuration file not found: {config_path}")
        sys.exit(1)

    config = load_config(config_path)

    # Initialize Zabbix API
    zabbix = ZabbixAPI(config["zabbix"]["url"], config["zabbix"]["token"])

    # Test connection
    try:
        api_version = zabbix._call("apiinfo.version", use_auth=False)
        print(f"Connected to Zabbix API version: {api_version}")
    except Exception as e:
        print(f"Error connecting to Zabbix: {e}")
        sys.exit(1)

    # Calculate date ranges
    date_calc = DateRangeCalculator()

    # Determine SLA period label
    if args.period == "day":
        sla_period_name = "1 Day"
    elif args.period == "week":
        sla_period_name = "7 Days"
    else:
        sla_period_name = "Previous Month"

    print(f"SLA Period: {sla_period_name}")

    # Get availability periods (relative to current date)
    availability_periods = date_calc.get_availability_periods()

    # Get default SLA settings
    default_sla = config.get("default_sla_threshold", 99.9)
    default_orange = config.get("default_orange_threshold", 5.0)
    report_mode = config.get("report_mode", "combined")

    # Get global excluded hosts
    global_excluded = config.get("global_excluded_hosts", []) or []
    global_excluded_lower = [h.lower() for h in global_excluded]
    if global_excluded:
        print(f"Globally excluding hosts: {', '.join(global_excluded)}")

    # Get host groups configuration
    host_groups_config = config.get("host_groups", {})

    # Handle command line override or legacy list format
    if args.groups:
        # Command line groups use default SLA
        host_groups_config = {g: {} for g in args.groups}
    elif isinstance(host_groups_config, list):
        # Legacy list format - convert to dict
        host_groups_config = {g: {} for g in host_groups_config}

    if not host_groups_config:
        print("Error: No host groups specified in config or command line")
        sys.exit(1)

    group_names = list(host_groups_config.keys())
    print(f"Processing host groups: {', '.join(group_names)}")
    print(f"Report mode: {report_mode}")

    # Fetch all host groups from Zabbix
    groups = zabbix.get_host_groups(group_names)

    if not groups:
        print(f"Warning: No host groups found matching: {group_names}")
        sys.exit(1)

    # For combined mode, create one report generator
    combined_report = None
    all_group_summaries = []

    if report_mode == "combined":
        combined_report = ExcelReportGenerator(default_sla, default_orange)

    output_config = config.get("output", {})
    output_dir = output_config.get("output_dir", ".")
    prefix = output_config.get("filename_prefix", "SLA_Report")
    include_timestamp = output_config.get("include_timestamp", True)

    for group in groups:
        group_name = group["name"]
        group_id = group["groupid"]

        # Get group-specific config
        group_config = host_groups_config.get(group_name, {})
        sla_threshold = group_config.get("sla_threshold", default_sla)
        orange_threshold = group_config.get("orange_threshold", default_orange)
        group_excluded = group_config.get("excluded_hosts", []) or []
        group_excluded_lower = [h.lower() for h in group_excluded]

        # Combine global and group-specific exclusions
        all_excluded_lower = global_excluded_lower + group_excluded_lower

        print(f"\nProcessing group: {group_name}")
        print(f"  SLA Threshold: {sla_threshold}%")
        if group_excluded:
            print(f"  Group-specific exclusions: {', '.join(group_excluded)}")

        # Get hosts in group
        hosts = zabbix.get_hosts_in_group(group_id)
        print(f"  Found {len(hosts)} hosts")

        if not hosts:
            continue

        host_data_list = []
        summary = {
            "group_name": group_name,
            "sla_threshold": sla_threshold,
            "total": 0,
            "compliant": 0,
            "warning": 0,
            "breach": 0,
        }

        for host in hosts:
            host_id = host["hostid"]
            host_name = host["name"]
            host_technical = host["host"]

            # Check if host should be excluded (global + group-specific)
            if host_name.lower() in all_excluded_lower or host_technical.lower() in all_excluded_lower:
                print(f"    Skipping (excluded): {host_name}")
                continue

            print(f"    Processing: {host_name}")

            # Calculate availability for each period
            avail_1_day = zabbix.get_host_availability(
                host_id,
                int(availability_periods["1_day"][0].timestamp()),
                int(availability_periods["1_day"][1].timestamp()),
            )

            avail_2_to_7 = zabbix.get_host_availability(
                host_id,
                int(availability_periods["7_days"][0].timestamp()),
                int(availability_periods["7_days"][1].timestamp()),
            )

            avail_2_to_30 = zabbix.get_host_availability(
                host_id,
                int(availability_periods["prev_month"][0].timestamp()),
                int(availability_periods["prev_month"][1].timestamp()),
            )

            # Device SLA is based on the selected period
            if args.period == "day":
                device_sla = avail_1_day["availability"]
            elif args.period == "week":
                device_sla = avail_2_to_7["availability"]
            else:  # month
                device_sla = avail_2_to_30["availability"]

            host_data = {
                "name": host_name,
                "host": host_technical,
                "avail_1_day": avail_1_day["availability"],
                "avail_7_days": avail_2_to_7["availability"],
                "avail_prev_month": avail_2_to_30["availability"],
                "device_sla": device_sla,
                # Store actual seconds for overall calculation
                "downtime_1_day": avail_1_day["downtime_seconds"],
                "downtime_7_days": avail_2_to_7["downtime_seconds"],
                "downtime_prev_month": avail_2_to_30["downtime_seconds"],
                "total_1_day": avail_1_day["total_seconds"],
                "total_7_days": avail_2_to_7["total_seconds"],
                "total_prev_month": avail_2_to_30["total_seconds"],
            }

            host_data_list.append(host_data)

            # Update summary counts
            summary["total"] += 1
            if device_sla >= sla_threshold:
                summary["compliant"] += 1
            elif device_sla >= sla_threshold - orange_threshold:
                summary["warning"] += 1
            else:
                summary["breach"] += 1

        # Calculate overall SLA for the group based on TOTAL TIME (not average)
        # Formula: ((Total Possible Uptime - Total Downtime) / Total Possible Uptime) × 100
        if host_data_list:
            # Sum all downtime and total time across all devices
            total_downtime_1_day = sum(h["downtime_1_day"] for h in host_data_list)
            total_downtime_7_days = sum(h["downtime_7_days"] for h in host_data_list)
            total_downtime_prev_month = sum(h["downtime_prev_month"] for h in host_data_list)

            total_possible_1_day = sum(h["total_1_day"] for h in host_data_list)
            total_possible_7_days = sum(h["total_7_days"] for h in host_data_list)
            total_possible_prev_month = sum(h["total_prev_month"] for h in host_data_list)

            # Calculate overall SLA based on total time
            overall_1_day = ((total_possible_1_day - total_downtime_1_day) / total_possible_1_day * 100) if total_possible_1_day > 0 else 100.0
            overall_7_days = ((total_possible_7_days - total_downtime_7_days) / total_possible_7_days * 100) if total_possible_7_days > 0 else 100.0
            overall_prev_month = ((total_possible_prev_month - total_downtime_prev_month) / total_possible_prev_month * 100) if total_possible_prev_month > 0 else 100.0

            # Overall SLA based on selected period
            if args.period == "day":
                overall_sla = overall_1_day
            elif args.period == "week":
                overall_sla = overall_7_days
            else:
                overall_sla = overall_prev_month

            summary["overall_sla"] = round(overall_sla, 2)
            summary["overall_1_day"] = round(overall_1_day, 2)
            summary["overall_7_days"] = round(overall_7_days, 2)
            summary["overall_prev_month"] = round(overall_prev_month, 2)
        else:
            summary["overall_sla"] = 100.0
            summary["overall_1_day"] = 100.0
            summary["overall_7_days"] = 100.0
            summary["overall_prev_month"] = 100.0

        all_group_summaries.append(summary)

        if report_mode == "separate":
            # Create separate report for this group
            report = ExcelReportGenerator(sla_threshold, orange_threshold)
            report.create_sheet(group_name, host_data_list, sla_threshold)
            report.add_summary_sheet([summary])

            # Generate filename for this group
            safe_group_name = group_name.replace(" ", "_").replace("/", "-")
            if include_timestamp:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{output_dir}/{prefix}_{safe_group_name}_{args.period}_{timestamp}.xlsx"
            else:
                output_file = f"{output_dir}/{prefix}_{safe_group_name}_{args.period}.xlsx"

            report.save(output_file)
            print(f"  Report saved: {output_file}")
        else:
            # Add to combined report
            combined_report.sla_threshold = sla_threshold
            combined_report.orange_threshold = orange_threshold
            combined_report.create_sheet(group_name, host_data_list, sla_threshold)

    # Save combined report if in combined mode
    if report_mode == "combined" and combined_report:
        combined_report.add_summary_sheet(all_group_summaries)

        if args.output:
            output_file = args.output
        else:
            if include_timestamp:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"{output_dir}/{prefix}_{args.period}_{timestamp}.xlsx"
            else:
                output_file = f"{output_dir}/{prefix}_{args.period}.xlsx"

        combined_report.save(output_file)

    print(f"\nReport generation complete!")
    print(f"\nSummary by group:")
    for summary in all_group_summaries:
        sla = summary.get("sla_threshold", default_sla)
        print(f"  {summary['group_name']}: SLA {sla}% | {summary['total']} hosts | {summary['compliant']} compliant | {summary['warning']} warning | {summary['breach']} breach")


if __name__ == "__main__":
    main()
